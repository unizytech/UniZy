"""
POC Metrics Service

Orchestrates the per-consultation tracker rows, per-day aggregate rows, and
per-recording timing tables for the Admin "POC Metrics" screen.

Design:
  - get_tracker_rows(...)   → list of dicts, one per consultation
  - get_aggregate_rows(...) → metrics × N day columns
  - get_timings_tables(...) → { 'doctor_all': [...], 'attendant_nurse': [...] }
  - build_xlsx(...)         → bytes of a single xlsx with 4 sheets
"""

import re
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook

from services.supabase_service import supabase
from services.edit_classifier import (
    classify_extraction_edits,
    classify_edit,
    count_date_errors,
    is_empty_like,
    DATE_FIELD_PATHS,
    _get_by_path,
    _normalize_value,
)

IST = timezone(timedelta(hours=5, minutes=30))

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_SPEAKER_PREFIX_RE = re.compile(r'^(Counsellor|Student|Speaker\s*\d+)\s*:', re.MULTILINE | re.IGNORECASE)


def _count_speakers(transcript: Optional[str]) -> int:
    """Count distinct Counsellor/Student/Speaker N prefixes in the transcript."""
    if not transcript:
        return 0
    prefixes = {m.group(1).strip().lower().replace(" ", "") for m in _SPEAKER_PREFIX_RE.finditer(transcript)}
    return len(prefixes)


def _short_id(uuid_str: Optional[str]) -> str:
    return uuid_str[:8] if uuid_str else ""


def _ist_date(dt_str: str) -> date:
    """Convert an ISO UTC timestamp string to an IST date."""
    dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    return dt.astimezone(IST).date()


def _ist_time(dt_str: str) -> str:
    dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    return dt.astimezone(IST).strftime("%H:%M:%S")


def _range_utc_bounds(start: date, end: date) -> Tuple[str, str]:
    """Convert an inclusive IST date range to UTC ISO bounds for filtering."""
    start_utc = datetime.combine(start, datetime.min.time(), tzinfo=IST).astimezone(timezone.utc)
    end_utc = datetime.combine(end + timedelta(days=1), datetime.min.time(), tzinfo=IST).astimezone(timezone.utc)
    return start_utc.isoformat(), end_utc.isoformat()


# ---------------------------------------------------------------------------
# Underlying data fetch (shared by tracker + aggregate + timings)
# ---------------------------------------------------------------------------


def _fetch_core_dataset(
    school_id: str,
    counsellor_id: Optional[str],
    assistant_id: Optional[str],
    start: date,
    end: date,
) -> List[Dict[str, Any]]:
    """Return a flat per-session dataset joined with job + extraction + metrics.

    Each dict has keys:
      session_id, created_at, counsellor_id, counsellor_name, assistant_id, template_code,
      total_duration_seconds, total_chunks, pj_status, pj_transcript,
      pj_error_message, pj_total_processing_time_seconds,
      pj_stitching_time_seconds, pj_transcription_time_seconds,
      pj_extraction_time_seconds,
      extraction_id, edit_count, original_extraction_json, edited_extraction_json,
      accuracy (dict with overall_wer, entity_errors) | None
    """
    start_iso, end_iso = _range_utc_bounds(start, end)

    # ---- Counsellors in this school (for name lookup + school filter) ----
    counsellors_res = supabase.table("counsellors")\
        .select("id, full_name, school_id")\
        .eq("school_id", school_id)\
        .execute()
    counsellors = counsellors_res.data or []
    counsellor_by_id = {d["id"]: d["full_name"] for d in counsellors}
    counsellor_ids_in_school = list(counsellor_by_id.keys())
    if not counsellor_ids_in_school:
        return []

    # ---- Assistants in this school (for name lookup when session is assistant-recorded) ----
    assistants_res = supabase.table("assistants")\
        .select("id, full_name, school_id")\
        .eq("school_id", school_id)\
        .execute()
    assistant_by_id = {n["id"]: n["full_name"] for n in (assistants_res.data or [])}

    # ---- Sessions matching filters ----
    # Exclude in-progress / cancelled — only consultations that actually submitted
    q = supabase.table("recording_sessions")\
        .select("id, created_at, counsellor_id, assistant_id, template_code, total_duration_seconds, total_chunks, status")\
        .gte("created_at", start_iso)\
        .lt("created_at", end_iso)\
        .not_.in_("status", ["RECORDING", "CANCELLED"])
    if counsellor_id:
        q = q.eq("counsellor_id", counsellor_id)
    else:
        q = q.in_("counsellor_id", counsellor_ids_in_school)
    if assistant_id:
        q = q.eq("assistant_id", assistant_id)
    q = q.order("created_at")
    sessions = q.execute().data or []
    session_ids = [s["id"] for s in sessions]
    if not session_ids:
        return []

    # ---- Latest processing_job per session ----
    pj_rows = supabase.table("processing_jobs")\
        .select("session_id, submission_id, status, transcript, error_message, "
                "stitching_time_seconds, transcription_time_seconds, "
                "extraction_time_seconds, total_processing_time_seconds, created_at")\
        .in_("session_id", session_ids)\
        .order("created_at", desc=True)\
        .execute().data or []
    latest_pj = {}
    for row in pj_rows:
        latest_pj.setdefault(row["session_id"], row)

    # ---- Latest medical_extraction per session ----
    me_rows = supabase.table("extractions")\
        .select("id, session_id, edit_count, original_extraction_json, "
                "edited_extraction_json, created_at")\
        .in_("session_id", session_ids)\
        .order("created_at", desc=True)\
        .execute().data or []
    latest_me = {}
    for row in me_rows:
        latest_me.setdefault(row["session_id"], row)

    # ---- Accuracy metrics for those extractions ----
    ext_ids = [me["id"] for me in latest_me.values()]
    accuracy_by_ext = {}
    if ext_ids:
        am_rows = supabase.table("extraction_accuracy_metrics")\
            .select("extraction_id, overall_wer, overall_wer_adjusted, overall_wer_adjusted_descriptions, entity_errors, total_words_ai_original, segment_metrics, computed_at")\
            .in_("extraction_id", ext_ids)\
            .order("computed_at", desc=True)\
            .execute().data or []
        for row in am_rows:
            accuracy_by_ext.setdefault(row["extraction_id"], row)

    # ---- Compose flat dataset ----
    dataset = []
    for s in sessions:
        pj = latest_pj.get(s["id"], {})
        me = latest_me.get(s["id"], {})
        acc = accuracy_by_ext.get(me.get("id")) if me else None
        # Name shown in the Tracker's "Name" column = whoever recorded the session
        _assistant_id = s.get("assistant_id")
        _counsellor_id = s.get("counsellor_id")
        if _assistant_id and _assistant_id in assistant_by_id:
            display_name = assistant_by_id[_assistant_id]
        else:
            display_name = counsellor_by_id.get(_counsellor_id, "")

        dataset.append({
            "session_id": s["id"],
            "created_at": s["created_at"],
            "counsellor_id": _counsellor_id,
            "counsellor_name": counsellor_by_id.get(_counsellor_id, ""),
            "assistant_id": _assistant_id,
            "display_name": display_name,
            "template_code": s.get("template_code") or "",
            "total_duration_seconds": float(s["total_duration_seconds"] or 0),
            "total_chunks": int(s.get("total_chunks") or 0),
            "pj_status": pj.get("status") or "",
            "pj_transcript": pj.get("transcript") or "",
            "pj_error_message": pj.get("error_message") or "",
            "pj_total_processing_time_seconds": float(pj["total_processing_time_seconds"]) if pj.get("total_processing_time_seconds") is not None else None,
            "pj_stitching_time_seconds": float(pj["stitching_time_seconds"]) if pj.get("stitching_time_seconds") is not None else None,
            "pj_transcription_time_seconds": float(pj["transcription_time_seconds"]) if pj.get("transcription_time_seconds") is not None else None,
            "pj_extraction_time_seconds": float(pj["extraction_time_seconds"]) if pj.get("extraction_time_seconds") is not None else None,
            "extraction_id": me.get("id"),
            "edit_count": int(me.get("edit_count") or 0),
            "original_extraction_json": me.get("original_extraction_json"),
            "edited_extraction_json": me.get("edited_extraction_json"),
            "accuracy": acc,
        })
    return dataset


# ---------------------------------------------------------------------------
# Tracker rows
# ---------------------------------------------------------------------------


def _entity_error_counts(accuracy: Optional[Dict[str, Any]]) -> Dict[str, int]:
    """Map entity_errors.by_type JSON to our four columns."""
    if not accuracy or not accuracy.get("entity_errors"):
        return {"rx": 0, "diagnosis": 0, "investigation": 0}
    by_type = (accuracy["entity_errors"] or {}).get("by_type") or {}
    drug = int(by_type.get("drug") or 0)
    dose = int(by_type.get("dose") or 0)
    duration = int(by_type.get("duration") or 0)
    diagnosis = int(by_type.get("diagnosis") or 0)
    lab_value = int(by_type.get("lab_value") or 0)
    return {
        "rx": drug + dose + duration,
        "diagnosis": diagnosis,
        "investigation": lab_value,
    }


def _entity_error_counts_live(rec: Dict[str, Any]) -> Dict[str, int]:
    """Re-compute entity error counts on the fly from raw JSON + transcript.

    Keeps the Tracker column values consistent with the detail modal and
    resilient to stale `extraction_accuracy_metrics` rows (e.g. when a
    counsellor's rows were computed before the JSON-string coercion or
    set-based matching landed).
    """
    from services.accuracy_metrics_service import _compute_entity_errors
    orig = _normalize_extraction(rec.get("original_extraction_json"))
    edit = _normalize_extraction(rec.get("edited_extraction_json"))
    if not orig or not edit:
        return {"rx": 0, "diagnosis": 0, "investigation": 0}
    transcript = rec.get("pj_transcript") or ""
    result = _compute_entity_errors(orig, edit, transcript)
    by_type = result.get("by_type") or {}
    return {
        "rx": int(by_type.get("drug") or 0) + int(by_type.get("dose") or 0) + int(by_type.get("duration") or 0),
        "diagnosis": int(by_type.get("diagnosis") or 0),
        "investigation": int(by_type.get("lab_value") or 0),
    }


def _row_for_session(rec: Dict[str, Any]) -> Dict[str, Any]:
    """Build one Tracker-sheet row from a core dataset record."""
    ist_dt = _ist_date(rec["created_at"])
    ist_t = _ist_time(rec["created_at"])
    audio_s = rec["total_duration_seconds"] or 0
    audio_min = round(audio_s / 60.0, 2) if audio_s else 0

    report_generated = rec["pj_status"] == "COMPLETED"
    report_gen_s = rec["pj_total_processing_time_seconds"]
    edited = rec["edit_count"] > 0

    edit_counts = {"major": 0, "minor": 0, "additive": 0, "dates": 0}
    if edited and rec.get("edited_extraction_json") and rec.get("original_extraction_json"):
        # Normalize JSON-string-encoded segments back to dict/list before classifying
        edit_counts = classify_extraction_edits(
            _normalize_extraction(rec["original_extraction_json"]),
            _normalize_extraction(rec["edited_extraction_json"]),
        )

    # Live-compute entity counts from raw JSON so Tracker column values
    # match the detail modal (and stay correct even when the stored
    # accuracy row is stale).
    entity_counts = (
        _entity_error_counts_live(rec) if edited
        else _entity_error_counts(rec.get("accuracy"))
    )
    speakers = _count_speakers(rec.get("pj_transcript"))

    return {
        "Date": ist_dt.isoformat(),
        "Name": rec.get("display_name") or rec.get("counsellor_name") or "",
        "Consult ID/No/Time": f"{_short_id(rec['session_id'])} @ {ist_t}",
        "Session ID": rec["session_id"],
        "Consult Duration (min)": audio_min,
        "No of Speakers": speakers,
        "Report Generated? (Y/N)": "Y" if report_generated else "N",
        "Error, if any": (rec.get("pj_error_message") or "")[:300],
        "Report Gen. Time (sec)": round(report_gen_s, 2) if report_gen_s is not None else None,
        "Edited? (Y/N)": "Y" if edited else "N",
        "Rx error, if any": entity_counts["rx"],
        "Diagnosis error, if any": entity_counts["diagnosis"],
        "Investigation error, if any": entity_counts["investigation"],
        "Dates error, if any": edit_counts["dates"],
        "Additive edits, if any": edit_counts["additive"],
        "Major edits, if any": edit_counts["major"],
        "Minor edits, if any": edit_counts["minor"],
        "Feedback, if any": "",
        "Comments, if any": "",
    }


TRACKER_COLS = [
    "Date", "Name", "Consult ID/No/Time", "Session ID", "Consult Duration (min)",
    "No of Speakers", "Report Generated? (Y/N)", "Error, if any",
    "Report Gen. Time (sec)", "Edited? (Y/N)", "Rx error, if any",
    "Diagnosis error, if any", "Investigation error, if any",
    "Dates error, if any", "Additive edits, if any", "Major edits, if any",
    "Minor edits, if any", "Feedback, if any", "Comments, if any",
]


def get_tracker_rows(
    school_id: str,
    counsellor_id: Optional[str],
    assistant_id: Optional[str],
    start: date,
    end: date,
) -> List[Dict[str, Any]]:
    dataset = _fetch_core_dataset(school_id, counsellor_id, assistant_id, start, end)
    return [_row_for_session(r) for r in dataset]


# ---------------------------------------------------------------------------
# Aggregate rows (day-by-day columns)
# ---------------------------------------------------------------------------


AGGREGATE_METRICS = [
    ("Throughput", "Total consultations"),
    ("Throughput", "Notes per counsellor"),
    ("Performance", "Avg consultation duration (min)"),
    ("Performance", "Avg report generation time (sec)"),
    ("Acceptance", "Signed off unchanged"),
    ("Edits", "Minor edits (count)"),
    ("Edits", "Major edits (count)"),
    ("Edits", "Additive edits (count)"),
    ("Medical Entity Errors", "Wrong drug / dose (count)"),
    ("Medical Entity Errors", "Wrong diagnosis (count)"),
    ("Medical Entity Errors", "Wrong lab value / units (count)"),
    ("Medical Entity Errors", "Wrong dates / duration (count)"),
    ("Medical Entity Errors", "Total medical entity errors"),
    ("Medical Entity Errors", "Medical Entity Error Rate (%)"),
    ("STT Accuracy", "WER (%)"),
    ("STT Accuracy", "Adjusted WER - paraphrasing (%)"),
    ("STT Accuracy", "Adjusted WER - description editing (%)"),
]


def _iter_dates(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)


def _count_entities_in_original(extraction_json: Dict[str, Any]) -> int:
    """Count list items across prescription, diagnosis, investigations (original only)."""
    if not isinstance(extraction_json, dict):
        return 0
    count = 0
    for key in ("prescription", "diagnosis", "investigations",
                "prescriptionOp", "diagnosisOp", "investigationsOp"):
        v = extraction_json.get(key)
        if isinstance(v, list):
            count += len(v)
    return count


def get_aggregate_rows(
    school_id: str,
    counsellor_id: Optional[str],
    assistant_id: Optional[str],
    start: date,
    end: date,
) -> Tuple[List[date], List[Dict[str, Any]]]:
    """Return (dates_in_range, aggregate_rows). Each row has 'category', 'metric',
    and one field per date keyed as 'YYYY-MM-DD'."""
    dataset = _fetch_core_dataset(school_id, counsellor_id, assistant_id, start, end)
    dates = list(_iter_dates(start, end))

    # Bucket records by IST date
    buckets: Dict[date, List[Dict[str, Any]]] = {d: [] for d in dates}
    for rec in dataset:
        d = _ist_date(rec["created_at"])
        if d in buckets:
            buckets[d].append(rec)

    rows = []
    for (category, metric) in AGGREGATE_METRICS:
        row = {"category": category, "metric": metric}
        for d in dates:
            key = d.isoformat()
            recs = buckets[d]
            total = len(recs)
            if metric == "Total consultations":
                row[key] = total
            elif metric == "Notes per counsellor":
                distinct_counsellors = len({r["counsellor_id"] for r in recs if r.get("counsellor_id")})
                row[key] = round(total / distinct_counsellors, 2) if distinct_counsellors else 0
            elif metric == "Avg consultation duration (min)":
                vals = [r["total_duration_seconds"] / 60.0 for r in recs if r["total_duration_seconds"]]
                row[key] = round(sum(vals) / len(vals), 2) if vals else 0
            elif metric == "Avg report generation time (sec)":
                vals = [r["pj_total_processing_time_seconds"] for r in recs if r.get("pj_total_processing_time_seconds") is not None]
                row[key] = round(sum(vals) / len(vals), 2) if vals else 0
            elif metric == "Signed off unchanged":
                row[key] = sum(1 for r in recs if r["edit_count"] == 0)
            elif metric == "Minor edits (count)":
                row[key] = sum(
                    classify_extraction_edits(
                        _normalize_extraction(r["original_extraction_json"]),
                        _normalize_extraction(r["edited_extraction_json"]),
                    ).get("minor", 0)
                    if r["edit_count"] > 0 and r.get("edited_extraction_json") else 0
                    for r in recs
                )
            elif metric == "Major edits (count)":
                row[key] = sum(
                    classify_extraction_edits(
                        _normalize_extraction(r["original_extraction_json"]),
                        _normalize_extraction(r["edited_extraction_json"]),
                    ).get("major", 0)
                    if r["edit_count"] > 0 and r.get("edited_extraction_json") else 0
                    for r in recs
                )
            elif metric == "Additive edits (count)":
                row[key] = sum(
                    classify_extraction_edits(
                        _normalize_extraction(r["original_extraction_json"]),
                        _normalize_extraction(r["edited_extraction_json"]),
                    ).get("additive", 0)
                    if r["edit_count"] > 0 and r.get("edited_extraction_json") else 0
                    for r in recs
                )
            elif metric == "Wrong drug / dose (count)":
                row[key] = sum(_entity_error_counts_live(r)["rx"] if r["edit_count"] > 0 else 0 for r in recs)
            elif metric == "Wrong diagnosis (count)":
                row[key] = sum(_entity_error_counts_live(r)["diagnosis"] if r["edit_count"] > 0 else 0 for r in recs)
            elif metric == "Wrong lab value / units (count)":
                row[key] = sum(_entity_error_counts_live(r)["investigation"] if r["edit_count"] > 0 else 0 for r in recs)
            elif metric == "Wrong dates / duration (count)":
                row[key] = sum(
                    count_date_errors(
                        _normalize_extraction(r["original_extraction_json"]),
                        _normalize_extraction(r["edited_extraction_json"]),
                    )
                    if r["edit_count"] > 0 and r.get("edited_extraction_json") else 0
                    for r in recs
                )
            elif metric == "Total medical entity errors":
                rx = sum(_entity_error_counts_live(r)["rx"] if r["edit_count"] > 0 else 0 for r in recs)
                dx = sum(_entity_error_counts_live(r)["diagnosis"] if r["edit_count"] > 0 else 0 for r in recs)
                inv = sum(_entity_error_counts_live(r)["investigation"] if r["edit_count"] > 0 else 0 for r in recs)
                dates_err = sum(
                    count_date_errors(
                        _normalize_extraction(r["original_extraction_json"]),
                        _normalize_extraction(r["edited_extraction_json"]),
                    )
                    if r["edit_count"] > 0 and r.get("edited_extraction_json") else 0
                    for r in recs
                )
                row[key] = rx + dx + inv + dates_err
            elif metric == "Medical Entity Error Rate (%)":
                total_entities = sum(_count_entities_in_original(r["original_extraction_json"]) for r in recs)
                rx = sum(_entity_error_counts_live(r)["rx"] if r["edit_count"] > 0 else 0 for r in recs)
                dx = sum(_entity_error_counts_live(r)["diagnosis"] if r["edit_count"] > 0 else 0 for r in recs)
                inv = sum(_entity_error_counts_live(r)["investigation"] if r["edit_count"] > 0 else 0 for r in recs)
                dates_err = sum(
                    count_date_errors(
                        _normalize_extraction(r["original_extraction_json"]),
                        _normalize_extraction(r["edited_extraction_json"]),
                    )
                    if r["edit_count"] > 0 and r.get("edited_extraction_json") else 0
                    for r in recs
                )
                total_errors = rx + dx + inv + dates_err
                row[key] = round((total_errors / total_entities) * 100.0, 2) if total_entities else 0
            elif metric == "WER (%)":
                # Pooled raw WER across extractions: Σ(errors) / Σ(AI words).
                num = 0.0
                den = 0
                for r in recs:
                    acc = r.get("accuracy")
                    if not acc or acc.get("overall_wer") is None:
                        continue
                    aw = int(acc.get("total_words_ai_original") or 0)
                    if aw <= 0:
                        continue
                    num += float(acc["overall_wer"]) * aw
                    den += aw
                row[key] = round((num / den) * 100.0, 2) if den else 0
            elif metric == "Adjusted WER - paraphrasing (%)":
                # Pooled WER after subtracting clinical-paraphrase matches
                # (hypoglycemia ↔ sugar, constipation ↔ motion, etc.).
                # Falls back to overall_wer when adjusted is missing.
                num = 0.0
                den = 0
                for r in recs:
                    acc = r.get("accuracy")
                    if not acc:
                        continue
                    wer_adj = acc.get("overall_wer_adjusted")
                    if wer_adj is None:
                        wer_adj = acc.get("overall_wer")
                    if wer_adj is None:
                        continue
                    aw = int(acc.get("total_words_ai_original") or 0)
                    if aw <= 0:
                        continue
                    num += float(wer_adj) * aw
                    den += aw
                row[key] = round((num / den) * 100.0, 2) if den else 0
            elif metric == "Adjusted WER - description editing (%)":
                # Pooled WER after subtracting paraphrases AND deletion errors.
                # Deletions in description-style free-text fields (e.g.
                # chiefComplaints[*].description) are typically the counsellor
                # trimming verbose AI prose, not real STT errors. Falls back
                # to the paraphrase-adjusted value (then to raw WER) when the
                # newer column hasn't been backfilled yet.
                num = 0.0
                den = 0
                for r in recs:
                    acc = r.get("accuracy")
                    if not acc:
                        continue
                    wer_desc = acc.get("overall_wer_adjusted_descriptions")
                    if wer_desc is None:
                        wer_desc = acc.get("overall_wer_adjusted")
                    if wer_desc is None:
                        wer_desc = acc.get("overall_wer")
                    if wer_desc is None:
                        continue
                    aw = int(acc.get("total_words_ai_original") or 0)
                    if aw <= 0:
                        continue
                    num += float(wer_desc) * aw
                    den += aw
                row[key] = round((num / den) * 100.0, 2) if den else 0
            else:
                row[key] = 0
        rows.append(row)

    return dates, rows


# ---------------------------------------------------------------------------
# Timings tables
# ---------------------------------------------------------------------------


TIMINGS_COLS = [
    "ist_time", "sess", "recorded_by", "template", "audio_s", "total_chunks",
    "pj_status", "stitch_s", "transcribe_s", "extract_s", "total_pipe_s",
    "pipe_to_audio_ratio", "error",
]


def _timings_row(rec: Dict[str, Any]) -> Dict[str, Any]:
    audio_s = rec["total_duration_seconds"] or 0
    total_pipe = rec.get("pj_total_processing_time_seconds")
    ratio = round(total_pipe / audio_s, 2) if (audio_s and total_pipe) else None
    return {
        "ist_time": _ist_time(rec["created_at"]),
        "sess": _short_id(rec["session_id"]),
        "recorded_by": "nurse" if rec.get("assistant_id") else "doctor",
        "template": rec["template_code"],
        "audio_s": round(audio_s, 2),
        "total_chunks": rec["total_chunks"],
        "pj_status": rec["pj_status"],
        "stitch_s": rec.get("pj_stitching_time_seconds"),
        "transcribe_s": rec.get("pj_transcription_time_seconds"),
        "extract_s": rec.get("pj_extraction_time_seconds"),
        "total_pipe_s": total_pipe,
        "pipe_to_audio_ratio": ratio,
        "error": (rec.get("pj_error_message") or "")[:200],
    }


def get_timings_tables(
    school_id: str,
    counsellor_id: Optional[str],
    assistant_id: Optional[str],
    start: date,
    end: date,
) -> Dict[str, List[Dict[str, Any]]]:
    """Returns {'doctor_all': rows-for-selected-counsellor, 'attendant_nurse': rows-for-selected-assistant}.

    - doctor_all: all sessions for the selected counsellor (counsellor-recorded + assistant-recorded)
    - attendant_nurse: sessions for the selected assistant; empty when assistant_id is None
    """
    counsellor_rows: List[Dict[str, Any]] = []
    if counsellor_id:
        ds = _fetch_core_dataset(school_id, counsellor_id, None, start, end)
        counsellor_rows = [_timings_row(r) for r in ds]

    assistant_rows: List[Dict[str, Any]] = []
    if assistant_id:
        ds = _fetch_core_dataset(school_id, None, assistant_id, start, end)
        assistant_rows = [_timings_row(r) for r in ds]

    return {"doctor_all": counsellor_rows, "attendant_nurse": assistant_rows}


# ---------------------------------------------------------------------------
# Metric detail — drill-down for clickable cells in Tracker / Aggregate
# ---------------------------------------------------------------------------


# Metric keys we support drill-down for. Aligns with Tracker columns and
# Aggregate metrics. Values here are canonical metric codes the UI passes
# as `?metric=<code>`.
METRIC_CODES = {
    "major_edits", "minor_edits", "additive_edits",
    "rx_error", "diagnosis_error", "investigation_error", "dates_error",
    "wer",
}


def _summarize(v: Any, limit: int = 160) -> str:
    """Compact string representation for a JSON value (for modal display)."""
    if v is None:
        return "—"
    if isinstance(v, str):
        s = v.strip()
    else:
        import json
        s = json.dumps(v, ensure_ascii=False)
    s = s.replace("\n", " ")
    return s if len(s) <= limit else s[: limit - 1] + "…"


def _scalar(v: Any) -> str:
    """Short, human-readable form of a scalar for diff display.

    Scalars (None/str/number/bool) render directly. Dict/list fall back to a
    compact bullet form (not raw JSON) so the modal stays readable.
    """
    if v is None or v == "" or v == [] or v == {}:
        return '""'
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, str):
        s = v.strip().replace("\n", " · ")
        return s if len(s) <= 80 else s[:79] + "…"
    if isinstance(v, dict):
        # Human-readable key=value summary, first 3 non-empty fields
        parts = []
        for k, val in v.items():
            if val in (None, "", [], {}):
                continue
            parts.append(f"{k}={_scalar(val) if not isinstance(val, (dict, list)) else '…'}")
            if len(parts) >= 3:
                break
        inside = ", ".join(parts) if parts else "(empty)"
        s = "{" + inside + "}"
        return s if len(s) <= 80 else s[:79] + "…"
    if isinstance(v, list):
        if not v:
            return "[]"
        parts = [_scalar(item) for item in v[:3]]
        inside = ", ".join(parts)
        suffix = f", …+{len(v) - 3}" if len(v) > 3 else ""
        s = "[" + inside + suffix + "]"
        return s if len(s) <= 80 else s[:79] + "…"
    return str(v)


def _diff_display(orig: Any, edited: Any, max_diffs: int = 8) -> List[str]:
    """Build concise 'orig → edited' lines, one per changed field/item.

    Output is human-readable (no raw JSON blobs). For list-of-dicts, diffs are
    emitted per item (e.g. "[0].name: CBC → CBC w/ diff"). Type mismatches
    (list→string, dict→list) are flagged explicitly.
    """
    lines: List[str] = []

    from services.edit_classifier import is_empty_like

    def _walk_dict(prefix: str, a: Dict[str, Any], b: Dict[str, Any]):
        keys = list(a.keys()) + [k for k in b.keys() if k not in a]
        for k in keys:
            if len(lines) >= max_diffs + 1:
                break
            av, bv = a.get(k), b.get(k)
            if av == bv:
                continue
            # Skip empty-key noise (missing in one side, empty in the other)
            if is_empty_like(av) and is_empty_like(bv):
                continue
            path = f"{prefix}{k}" if prefix else k
            if isinstance(av, dict) and isinstance(bv, dict):
                _walk_dict(f"{path}.", av, bv)
            elif isinstance(av, list) and isinstance(bv, list):
                _walk_list(path, av, bv)
            else:
                lines.append(f"{path}: {_scalar(av)} → {_scalar(bv)}")

    def _walk_list(prefix: str, a: List[Any], b: List[Any]):
        for i in range(max(len(a), len(b))):
            if len(lines) >= max_diffs + 1:
                return
            o = a[i] if i < len(a) else None
            e = b[i] if i < len(b) else None
            if o == e:
                continue
            if is_empty_like(o) and is_empty_like(e):
                continue
            idx_prefix = f"{prefix}[{i}]" if prefix else f"[{i}]"
            if isinstance(o, dict) and isinstance(e, dict):
                _walk_dict(f"{idx_prefix}.", o, e)
            elif o is None:
                lines.append(f"{idx_prefix}: (added) → {_scalar(e)}")
            elif e is None:
                lines.append(f"{idx_prefix}: {_scalar(o)} → (removed)")
            else:
                lines.append(f"{idx_prefix}: {_scalar(o)} → {_scalar(e)}")

    # Dispatch based on top-level types
    if isinstance(orig, dict) and isinstance(edited, dict):
        _walk_dict("", orig, edited)
    elif isinstance(orig, list) and isinstance(edited, list):
        _walk_list("", orig, edited)
    elif type(orig) is not type(edited) and not (orig is None or edited is None):
        # Type mismatch (e.g. list ↔ string): flag explicitly
        lines.append(f"[{type(orig).__name__}] → [{type(edited).__name__}]")
        lines.append(f"  {_scalar(orig)}")
        lines.append(f"  → {_scalar(edited)}")
    else:
        lines.append(f"{_scalar(orig)} → {_scalar(edited)}")

    if not lines:
        # Fallback when both values are equivalent at this heuristic depth
        lines.append(f"{_scalar(orig)} → {_scalar(edited)}")

    if len(lines) > max_diffs:
        extra = len(lines) - max_diffs
        lines = lines[:max_diffs] + [f"… +{extra} more"]
    return lines


def _normalize_extraction(doc: Any) -> Dict[str, Any]:
    """Defensive re-parse: if any top-level value is a JSON-encoded string
    (known EHR save-path artifact for list segments), parse it back."""
    from services.accuracy_metrics_service import _coerce_json
    if not isinstance(doc, dict):
        return {}
    return {k: _coerce_json(v) for k, v in doc.items()}


def _edit_detail_rows(rec: Dict[str, Any], magnitude_filter: Optional[str], additive_filter: bool) -> List[Dict[str, Any]]:
    """For one extraction, walk segments and emit one row per matching edit."""
    orig = _normalize_extraction(rec.get("original_extraction_json"))
    edit = _normalize_extraction(rec.get("edited_extraction_json"))
    if not isinstance(orig, dict) or not isinstance(edit, dict):
        return []
    out: List[Dict[str, Any]] = []
    seg_codes = set(orig.keys()) | set(edit.keys())
    for seg in sorted(seg_codes):
        # Same artifact guard classify_extraction_edits uses
        if seg in orig and seg not in edit:
            continue
        r = classify_edit(seg, orig.get(seg), edit.get(seg))
        if not r["changed"]:
            continue
        if additive_filter and not r["additive"]:
            continue
        if magnitude_filter and r["magnitude"] != magnitude_filter:
            continue
        out.append({
            "session_id": rec["session_id"],
            "name": rec.get("display_name") or rec.get("counsellor_name") or "",
            "segment": seg,
            "magnitude": r["magnitude"],
            "additive": r["additive"],
            "word_change_pct": r["word_change_pct"],
            "diff": _diff_display(orig.get(seg), edit.get(seg)),
        })
    return out


def _dates_detail_rows(rec: Dict[str, Any]) -> List[Dict[str, Any]]:
    orig = _normalize_extraction(rec.get("original_extraction_json"))
    edit = _normalize_extraction(rec.get("edited_extraction_json"))
    if not isinstance(orig, dict) or not isinstance(edit, dict):
        return []
    artifact_top = set(orig.keys()) - set(edit.keys())
    out: List[Dict[str, Any]] = []
    for path in DATE_FIELD_PATHS:
        top = path.split(".", 1)[0].split("[", 1)[0]
        if top in artifact_top:
            continue
        o = _get_by_path(orig, path)
        e = _get_by_path(edit, path)
        if isinstance(o, list) or isinstance(e, list):
            o_list = o or []
            e_list = e or []
            for idx in range(min(len(o_list), len(e_list))):
                if _normalize_value(o_list[idx]) != _normalize_value(e_list[idx]):
                    out.append({
                        "session_id": rec["session_id"],
                        "name": rec.get("display_name") or rec.get("counsellor_name") or "",
                        "segment": f"{path}[{idx}]",
                        "diff": [f"{_scalar(o_list[idx])} → {_scalar(e_list[idx])}"],
                    })
        else:
            if _normalize_value(o) != _normalize_value(e):
                out.append({
                    "session_id": rec["session_id"],
                    "name": rec.get("display_name") or rec.get("counsellor_name") or "",
                    "segment": path,
                    "diff": [f"{_scalar(o)} → {_scalar(e)}"],
                })
    return out


def _entity_detail_rows(rec: Dict[str, Any], kind: str) -> List[Dict[str, Any]]:
    """Drill-down for Rx / Diagnosis / Investigation entity errors.

    kind ∈ {"rx_error", "diagnosis_error", "investigation_error"}. Re-runs the
    set-based matching so it stays aligned with the stored by_type counts.
    """
    from services.accuracy_metrics_service import (
        _match_items, _drug_name, _inv_name, _dx_key, _dx_display_name,
        _key_in_transcript, _tokenize,
    )
    orig = _normalize_extraction(rec.get("original_extraction_json"))
    edit = _normalize_extraction(rec.get("edited_extraction_json"))
    transcript = rec.get("pj_transcript") or ""
    transcript_tokens = set(_tokenize(transcript))

    keys_for_kind = {
        "rx_error": (["prescription", "medications", "drugs", "rx"], _drug_name, lambda it: (it.get("name") if isinstance(it, dict) else str(it)) or ""),
        "investigation_error": (["investigations", "lab_tests", "diagnostics"], _inv_name, lambda it: (it.get("name") if isinstance(it, dict) else str(it)) or ""),
        "diagnosis_error": (["diagnosis", "diagnoses", "provisional_diagnosis", "final_diagnosis"], _dx_key, _dx_display_name),
    }
    seg_keys, key_fn, display_fn = keys_for_kind[kind]

    out: List[Dict[str, Any]] = []
    for k in seg_keys:
        if k in orig and k not in edit:
            continue
        ai_items = orig.get(k, [])
        doc_items = edit.get(k, [])
        if not isinstance(ai_items, list) or not isinstance(doc_items, list):
            continue
        matched, ai_only, doc_only = _match_items(ai_items, doc_items, key_fn)
        # Dose-only substitutions in matched pairs (Rx only)
        if kind == "rx_error":
            for ai_i, doc_i in matched:
                if not isinstance(ai_i, dict) or not isinstance(doc_i, dict):
                    continue
                ai_dose = str(ai_i.get("dose", ai_i.get("dosage", ""))).strip().lower()
                doc_dose = str(doc_i.get("dose", doc_i.get("dosage", ""))).strip().lower()
                if ai_dose and doc_dose and ai_dose != doc_dose and not _key_in_transcript(ai_dose, transcript_tokens):
                    out.append({
                        "session_id": rec["session_id"],
                        "name": rec.get("display_name") or rec.get("counsellor_name") or "",
                        "segment": k,
                        "kind": "dose_changed",
                        "item": display_fn(ai_i),
                        "diff": [f"dose: {_scalar(ai_dose)} → {_scalar(doc_dose)}"],
                    })
        for ai_i in ai_only:
            name = key_fn(ai_i)
            if not _key_in_transcript(name, transcript_tokens):
                out.append({
                    "session_id": rec["session_id"],
                    "name": rec.get("display_name") or rec.get("counsellor_name") or "",
                    "segment": k,
                    "kind": "ai_hallucinated_removed",
                    "item": display_fn(ai_i) or name or "",
                    "diff": [f"{_scalar(display_fn(ai_i) or name)} → (removed)"],
                })
        for doc_i in doc_only:
            name = key_fn(doc_i)
            if _key_in_transcript(name, transcript_tokens):
                out.append({
                    "session_id": rec["session_id"],
                    "name": rec.get("display_name") or rec.get("counsellor_name") or "",
                    "segment": k,
                    "kind": "ai_missed_added",
                    "item": display_fn(doc_i) or name or "",
                    "diff": [f"(not extracted) → {_scalar(display_fn(doc_i) or name)}"],
                })
    return out


def _wer_detail_rows(rec: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Per-segment WER breakdown for one extraction, read from stored
    accuracy_metrics.segment_metrics. Only segments with errors are returned.
    """
    acc = rec.get("accuracy") or {}
    seg_metrics = (acc.get("segment_metrics") or []) if isinstance(acc, dict) else []
    out: List[Dict[str, Any]] = []
    for s in seg_metrics:
        if not isinstance(s, dict):
            continue
        errs = (
            int(s.get("substitutions_ai_error") or 0)
            + int(s.get("deletions_ai_error") or 0)
            + int(s.get("insertions_ai_error") or 0)
        )
        if errs <= 0:
            continue
        out.append({
            "session_id": rec["session_id"],
            "name": rec.get("display_name") or rec.get("counsellor_name") or "",
            "segment": s.get("segment_code", ""),
            "errors": errs,
            "subs_ai_error": int(s.get("substitutions_ai_error") or 0),
            "dels_ai_error": int(s.get("deletions_ai_error") or 0),
            "ins_ai_error": int(s.get("insertions_ai_error") or 0),
            "ai_word_count": int(s.get("ai_word_count") or 0),
            "wer": float(s.get("wer") or 0.0),
        })
    # Sort by errors descending for quick scanning
    out.sort(key=lambda r: r["errors"], reverse=True)
    return out


def get_metric_detail(
    school_id: str,
    counsellor_id: Optional[str],
    assistant_id: Optional[str],
    start: date,
    end: date,
    metric: str,
    session_id: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Drill-down rows for a clicked Tracker or Aggregate cell.

    When `session_id` is set the result is scoped to that one session (Tracker
    cell). Otherwise it spans all sessions in the filter (Aggregate cell).
    """
    if metric not in METRIC_CODES:
        raise ValueError(f"Unknown metric: {metric}")

    dataset = _fetch_core_dataset(school_id, counsellor_id, assistant_id, start, end)
    if session_id:
        dataset = [r for r in dataset if r["session_id"] == session_id]

    rows: List[Dict[str, Any]] = []
    for rec in dataset:
        if rec.get("edit_count", 0) == 0 and metric != "wer":
            continue
        if metric == "major_edits":
            rows.extend(_edit_detail_rows(rec, "major", False))
        elif metric == "minor_edits":
            rows.extend(_edit_detail_rows(rec, "minor", False))
        elif metric == "additive_edits":
            rows.extend(_edit_detail_rows(rec, None, True))
        elif metric == "dates_error":
            rows.extend(_dates_detail_rows(rec))
        elif metric in {"rx_error", "diagnosis_error", "investigation_error"}:
            rows.extend(_entity_detail_rows(rec, metric))
        elif metric == "wer":
            rows.extend(_wer_detail_rows(rec))

    return rows


# ---------------------------------------------------------------------------
# XLSX export (4 sheets)
# ---------------------------------------------------------------------------


def _write_sheet(ws, columns: List[str], rows: List[Dict[str, Any]]):
    for c, col in enumerate(columns, 1):
        ws.cell(row=1, column=c, value=col)
    for r, row in enumerate(rows, 2):
        for c, col in enumerate(columns, 1):
            ws.cell(row=r, column=c, value=row.get(col))
    # Column widths
    for c, col in enumerate(columns, 1):
        max_len = max([len(str(col))] + [len(str(row.get(col) or "")) for row in rows])
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(max_len + 2, 60)


def build_xlsx(
    school_id: str,
    counsellor_id: Optional[str],
    assistant_id: Optional[str],
    start: date,
    end: date,
) -> bytes:
    """Build a 2-sheet Excel export: Tracker + Aggregate.

    Doctor_All and Attendant_Nurse timing tables are shown in the UI for
    drill-down but intentionally NOT exported to Excel — the Excel is
    intended to be the POC tracker artifact, not the ops-timing view.
    """
    tracker_rows = get_tracker_rows(school_id, counsellor_id, assistant_id, start, end)
    dates, agg_rows = get_aggregate_rows(school_id, counsellor_id, assistant_id, start, end)

    wb = Workbook()

    # Tracker
    ws1 = wb.active
    ws1.title = "Tracker"
    _write_sheet(ws1, TRACKER_COLS, tracker_rows)

    # Aggregate (category, metric, date columns)
    ws2 = wb.create_sheet("Aggregate")
    agg_cols = ["category", "metric"] + [d.isoformat() for d in dates]
    _write_sheet(ws2, agg_cols, agg_rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
